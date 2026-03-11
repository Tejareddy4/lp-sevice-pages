from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Iterable, Optional

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "csv" / "Service-pages.xlsx"
OUTPUT_ROOT = BASE_DIR / "service"


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def slugify(value: str) -> str:
    text = (value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-+", "-", text).strip("-")


def build_header_index(headers: Iterable[object]) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for i, header in enumerate(headers):
        key = normalize_header(header)
        if key:
            index[key] = i
    return index


def find_col(header_index: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for alias in aliases:
        normalized = normalize_header(alias)
        if normalized in header_index:
            return header_index[normalized]
    return None


def cell_text(row: tuple[object, ...], col_idx: Optional[int]) -> str:
    if col_idx is None or col_idx >= len(row):
        return ""
    value = row[col_idx]
    return "" if value is None else str(value)


def nowdoc(var_name: str, token: str, content: str) -> str:
    body = str(content or "").replace("\r\n", "\n").replace("\r", "\n")
    return f"${var_name} = <<<'{token}'\n{body}\n{token};"


def make_php_page(origin: str, destination: str, mapped: Dict[str, str]) -> str:
    lines = [
        "<?php",
        "",
        f'$Origin = "{origin}"; //city name',
        f'$Destination = "{destination}"; //country name',
        nowdoc("Excerpt", "EXCERPT", mapped["Excerpt"]),
        nowdoc("Updatedhtml", "UPDATEDHTML", mapped["Updatedhtml"]),
        nowdoc("Relposts", "RELPOSTS", mapped["Relposts"]),
        nowdoc("FAQ", "FAQ", mapped["FAQ"]),
        "",
        "//scripts",
        nowdoc("FAQscript", "FAQSCRIPT", mapped["FAQscript"]),
        nowdoc("Review_Schema", "REVIEWSCHEMA", mapped["Review_Schema"]),
        "",
        'include __DIR__ . "/../../template.php";',
        "",
        "?>",
        "",
    ]
    return "\n".join(lines)


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active

    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError("Sheet1 has no header row.")

    header_index = build_header_index(headers)

    # Explicit mapping requested:
    # $Origin = City, $Destination = Destination
    # Current workbook uses "Country" for destination in some sheets.
    city_col = find_col(header_index, ["City"])
    destination_col = find_col(header_index, ["Destination", "Country"])

    # Exact mapping requested by the user.
    excerpt_col = find_col(header_index, ["Excerpt"])
    updated_html_col = find_col(header_index, ["Updated Html content"])
    related_posts_col = find_col(header_index, ["Related Posts"])
    coded_faq_col = find_col(header_index, ["Coded FAQ"])
    faq_schema_col = find_col(header_index, ["FAQ Schema"])
    review_schema_col = find_col(header_index, ["Review_Schema"])

    missing = []
    required = {
        "City": city_col is not None,
        "Destination/Country": destination_col is not None,
        "Excerpt": excerpt_col is not None,
        "Updated Html content": updated_html_col is not None,
        "Related Posts": related_posts_col is not None,
        "Coded FAQ": coded_faq_col is not None,
        "FAQ Schema": faq_schema_col is not None,
        "Review_Schema": review_schema_col is not None,
    }
    for name, ok in required.items():
        if not ok:
            missing.append(name)
    if missing:
        raise ValueError(f"Missing required columns in Sheet1: {', '.join(missing)}")

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    created = 0

    for row in rows:
        city = cell_text(row, city_col).strip()
        destination = cell_text(row, destination_col).strip()
        if not city or not destination:
            continue

        city_slug = slugify(city)
        destination_slug = slugify(destination)
        if not city_slug or not destination_slug:
            continue

        page_name = f"{city_slug}-to-{destination_slug}-courier.php"
        city_dir = OUTPUT_ROOT / city_slug
        city_dir.mkdir(parents=True, exist_ok=True)

        mapped_values = {
            "Excerpt": cell_text(row, excerpt_col),
            "Updatedhtml": cell_text(row, updated_html_col),
            "Relposts": cell_text(row, related_posts_col),
            "FAQ": cell_text(row, coded_faq_col),
            "FAQscript": cell_text(row, faq_schema_col),
            "Review_Schema": cell_text(row, review_schema_col),
        }

        php_content = make_php_page(city, destination, mapped_values)
        (city_dir / page_name).write_text(php_content, encoding="utf-8")
        created += 1

    print(f"Created/updated {created} page file(s) under: {OUTPUT_ROOT}")


if __name__ == "__main__":
    main()
